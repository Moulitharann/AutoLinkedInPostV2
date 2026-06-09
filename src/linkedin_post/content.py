import base64
import csv
import hashlib
import io
import json
import os
import re
from typing import Any
from urllib.parse import parse_qs, urlparse

from openpyxl import load_workbook

from .http_utils import get_session
from .settings import Settings, require


class AllRowsPostedError(Exception):
    """Raised when the content source has no unused rows left."""


def download_url(source_url: str) -> str:
    parsed = urlparse(source_url)
    if parsed.netloc.endswith("docs.google.com") and "/spreadsheets/d/" in parsed.path:
        sheet_id = parsed.path.split("/spreadsheets/d/", 1)[1].split("/", 1)[0]
        params = parse_qs(parsed.query)
        gid = params.get("gid", ["0"])[0]
        return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
    if parsed.netloc.endswith("1drv.ms") or parsed.netloc.endswith("onedrive.live.com"):
        encoded_url = base64.urlsafe_b64encode(source_url.encode("utf-8")).decode("utf-8").rstrip("=")
        return f"https://api.onedrive.com/v1.0/shares/u!{encoded_url}/root/content"
    return source_url


def normalize_rows(rows: Any) -> list[dict[str, str]]:
    normalized = []
    for row in rows:
        title_key = next((k for k in row.keys() if k.lower() == "title"), None)
        desc_key = next((k for k in row.keys() if k.lower() == "description"), None)
        if not title_key or not desc_key:
            continue
        title = str(row.get(title_key, "")).strip()
        description = str(row.get(desc_key, "")).strip()
        if title and description:
            normalized.append({"title": title, "description": description})
    return normalized


def load_content_rows(settings: Settings) -> list[dict[str, str]]:
    source_url = download_url(require(settings.content_source_url, "CONTENT_SOURCE_URL"))
    response = get_session().get(source_url, timeout=60)
    if response.status_code in {401, 403} and "onedrive" in source_url:
        raise SystemExit(
            "OneDrive blocked the workbook download. Open the Excel file, choose Share, "
            "set access to 'Anyone with the link can view', then paste the new link into CONTENT_SOURCE_URL."
        )
    response.raise_for_status()

    source_type = settings.content_source_type
    if source_type == "auto":
        content_type = response.headers.get("Content-Type", "").lower()
        source_path = urlparse(source_url).path.lower()
        if source_path.endswith(".csv") or "csv" in content_type:
            source_type = "csv"
        else:
            source_type = "xlsx"

    if source_type == "csv":
        text = response.content.decode("utf-8-sig")
        return normalize_rows(csv.DictReader(io.StringIO(text)))

    if source_type == "xlsx":
        workbook = load_workbook(io.BytesIO(response.content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value or "").strip().lower() for value in rows[0]]
        parsed_rows = []
        for row in rows[1:]:
            parsed_rows.append(
                {
                    headers[index]: str(value or "").strip()
                    for index, value in enumerate(row)
                    if index < len(headers) and headers[index]
                }
            )
        return normalize_rows(parsed_rows)

    raise SystemExit("CONTENT_SOURCE_TYPE must be auto, csv, or xlsx.")


def load_history(settings: Settings) -> dict[str, Any]:
    if not os.path.exists(settings.post_history_file):
        return {"posted": []}
    with open(settings.post_history_file, "r", encoding="utf-8") as file:
        return json.load(file)


def save_history(settings: Settings, history: dict[str, Any]) -> None:
    with open(settings.post_history_file, "w", encoding="utf-8") as file:
        json.dump(history, file, indent=2)


def load_scheduler_state(settings: Settings) -> dict[str, Any]:
    if not os.path.exists(settings.scheduler_state_file):
        return {}
    with open(settings.scheduler_state_file, "r", encoding="utf-8") as file:
        return json.load(file)


def save_scheduler_state(settings: Settings, state: dict[str, Any]) -> None:
    with open(settings.scheduler_state_file, "w", encoding="utf-8") as file:
        json.dump(state, file, indent=2)


def row_key(row: dict[str, str]) -> str:
    content = f"{row['title']}\n{row['description']}"
    return hashlib.sha256(content.encode("utf-8")).hexdigest()


def next_content_row(settings: Settings) -> tuple[dict[str, str], str]:
    rows = load_content_rows(settings)
    if not rows:
        raise SystemExit("No usable rows found. Your source must have title and description columns.")

    history = load_history(settings)
    posted = set(history.get("posted", []))
    for row in rows:
        key = row_key(row)
        if key not in posted:
            return row, key

    raise AllRowsPostedError("All rows from the source have already been posted.")


def mark_posted(settings: Settings, key: str) -> None:
    history = load_history(settings)
    posted = history.setdefault("posted", [])
    if key not in posted:
        posted.append(key)
    save_history(settings, history)
