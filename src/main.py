import argparse
import base64
import csv
import hashlib
import io
import json
import os
import secrets
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime
from http.server import BaseHTTPRequestHandler, HTTPServer
from typing import Any
from urllib.parse import parse_qs, urlencode, urlparse
from zoneinfo import ZoneInfo

from requests.adapters import HTTPAdapter, Retry
import requests
from dotenv import load_dotenv
from openpyxl import load_workbook


LINKEDIN_AUTH_URL = "https://www.linkedin.com/oauth/v2/authorization"
LINKEDIN_TOKEN_URL = "https://www.linkedin.com/oauth/v2/accessToken"
LINKEDIN_USERINFO_URL = "https://api.linkedin.com/v2/userinfo"
LINKEDIN_POSTS_URL = "https://api.linkedin.com/rest/posts"
LINKEDIN_VERSION = "202605"


@dataclass(frozen=True)
class Settings:
    gemini_api_key: str
    gemini_model: str
    hugging_face_api_key: str
    linkedin_client_id: str
    linkedin_client_secret: str
    linkedin_redirect_uri: str
    linkedin_oauth_scopes: str
    linkedin_access_token: str
    linkedin_person_urn: str
    content_source_url: str
    content_source_type: str
    post_history_file: str
    scheduler_state_file: str
    post_interval_days: int
    post_topic: str
    post_tone: str
    post_audience: str
    post_visibility: str


def get_session() -> requests.Session:
    """Create a requests session with retry logic."""
    session = requests.Session()
    retries = Retry(
        total=3,
        backoff_factor=1,
        status_forcelist=[429, 500, 502, 503, 504],
    )
    adapter = HTTPAdapter(max_retries=retries)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    return session

def load_settings() -> Settings:
    load_dotenv()
    return Settings(
        gemini_api_key=os.getenv("GEMINI_API_KEY", ""),
        gemini_model=os.getenv("GEMINI_MODEL", "gemini-flash-latest"),
        hugging_face_api_key=os.getenv("HUGGING_FACE_API_KEY", ""),
        linkedin_client_id=os.getenv("LINKEDIN_CLIENT_ID", ""),
        linkedin_client_secret=os.getenv("LINKEDIN_CLIENT_SECRET", ""),
        linkedin_redirect_uri=os.getenv("LINKEDIN_REDIRECT_URI", "http://localhost:8000/callback"),
        linkedin_oauth_scopes=os.getenv("LINKEDIN_OAUTH_SCOPES", "openid profile w_member_social"),
        linkedin_access_token=os.getenv("LINKEDIN_ACCESS_TOKEN", ""),
        linkedin_person_urn=os.getenv("LINKEDIN_PERSON_URN", ""),
        content_source_url=os.getenv("CONTENT_SOURCE_URL", ""),
        content_source_type=os.getenv("CONTENT_SOURCE_TYPE", "auto").lower(),
        post_history_file=os.getenv("POST_HISTORY_FILE", "posts_history.json"),
        scheduler_state_file=os.getenv("SCHEDULER_STATE_FILE", "scheduler_state.json"),
        post_interval_days=int(os.getenv("POST_INTERVAL_DAYS", "2")),
        post_topic=os.getenv("POST_TOPIC", "software engineering"),
        post_tone=os.getenv("POST_TONE", "clear, practical, senior software engineer"),
        post_audience=os.getenv("POST_AUDIENCE", "software engineers"),
        post_visibility=os.getenv("POST_VISIBILITY", "PUBLIC"),
    )


def require(value: str, name: str) -> str:
    if not value:
        raise SystemExit(f"Missing {name}. Add it to .env and try again.")
    return value


def build_auth_url(settings: Settings) -> str:
    require(settings.linkedin_client_id, "LINKEDIN_CLIENT_ID")
    state = secrets.token_urlsafe(24)
    return build_auth_url_with_state(settings, state)


def build_auth_url_with_state(settings: Settings, state: str) -> str:
    query = urlencode(
        {
            "response_type": "code",
            "client_id": settings.linkedin_client_id,
            "redirect_uri": settings.linkedin_redirect_uri,
            "scope": settings.linkedin_oauth_scopes,
            "state": state,
        }
    )
    return f"{LINKEDIN_AUTH_URL}?{query}"


def extract_code(callback_url: str) -> str:
    parsed = urlparse(callback_url)
    params = parse_qs(parsed.query)
    error = params.get("error", [""])[0]
    if error:
        description = params.get("error_description", [""])[0]
        raise SystemExit(f"LinkedIn returned error: {error} {description}".strip())
    code = params.get("code", [""])[0]
    if not code:
        raise SystemExit("Could not find ?code= in the callback URL.")
    return code


def exchange_code_for_token(settings: Settings, callback_url: str) -> tuple[str, str]:
    require(settings.linkedin_client_id, "LINKEDIN_CLIENT_ID")
    require(settings.linkedin_client_secret, "LINKEDIN_CLIENT_SECRET")

    response = get_session().post(
        LINKEDIN_TOKEN_URL,
        data={
            "grant_type": "authorization_code",
            "code": extract_code(callback_url),
            "redirect_uri": settings.linkedin_redirect_uri,
            "client_id": settings.linkedin_client_id,
            "client_secret": settings.linkedin_client_secret,
        },
        timeout=30,
    )
    response.raise_for_status()
    token_data = response.json()
    access_token = token_data["access_token"]
    person_urn = person_urn_from_token_data(token_data, access_token)
    return access_token, person_urn


def exchange_raw_code_for_token(settings: Settings, code: str) -> tuple[str, str]:
    require(settings.linkedin_client_id, "LINKEDIN_CLIENT_ID")
    require(settings.linkedin_client_secret, "LINKEDIN_CLIENT_SECRET")

    response = get_session().post(
        LINKEDIN_TOKEN_URL,
        data={
            "grant_type": "authorization_code",
            "code": code,
            "redirect_uri": settings.linkedin_redirect_uri,
            "client_id": settings.linkedin_client_id,
            "client_secret": settings.linkedin_client_secret,
        },
        timeout=30,
    )
    response.raise_for_status()
    token_data = response.json()
    access_token = token_data["access_token"]
    person_urn = person_urn_from_token_data(token_data, access_token)
    return access_token, person_urn


def person_urn_from_token_data(token_data: dict[str, Any], access_token: str) -> str:
    id_token = token_data.get("id_token", "")
    if id_token:
        subject = subject_from_id_token(id_token)
        if subject:
            return f"urn:li:person:{subject}"
    if "openid" not in token_data.get("scope", ""):
        return ""
    return fetch_person_urn(access_token)


def subject_from_id_token(id_token: str) -> str:
    try:
        payload = id_token.split(".")[1]
        padded_payload = payload + "=" * (-len(payload) % 4)
        decoded = base64.urlsafe_b64decode(padded_payload.encode("utf-8"))
        claims = json.loads(decoded)
        return claims.get("sub", "")
    except (IndexError, ValueError, json.JSONDecodeError):
        return ""


def login(settings: Settings) -> None:
    parsed_redirect = urlparse(settings.linkedin_redirect_uri)
    if parsed_redirect.hostname != "localhost":
        raise SystemExit("The login command requires LINKEDIN_REDIRECT_URI to use localhost.")

    port = parsed_redirect.port or 80
    path = parsed_redirect.path or "/"
    state = secrets.token_urlsafe(24)
    result: dict[str, str] = {}

    class CallbackHandler(BaseHTTPRequestHandler):
        def do_GET(self) -> None:
            parsed = urlparse(self.path)
            params = parse_qs(parsed.query)

            if parsed.path != path:
                self.send_response(404)
                self.end_headers()
                return

            if params.get("state", [""])[0] != state:
                result["error"] = "OAuth state did not match."
            elif params.get("error", [""])[0]:
                result["error"] = params.get("error_description", params.get("error", ["Unknown error"]))[0]
            else:
                result["code"] = params.get("code", [""])[0]

            has_error = bool(result.get("error"))
            self.send_response(400 if has_error else 200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.end_headers()
            if has_error:
                message = result["error"].encode("utf-8", errors="replace")
                self.wfile.write(
                    b"<html><body><h2>LinkedIn authorization failed.</h2><p>"
                    + message
                    + b"</p><p>You can close this tab and return to the terminal.</p></body></html>"
                )
            else:
                self.wfile.write(
                    b"<html><body><h2>LinkedIn authorization captured.</h2>"
                    b"<p>You can close this tab and return to the terminal.</p></body></html>"
                )

        def log_message(self, format: str, *args: Any) -> None:
            return

    auth_url = build_auth_url_with_state(settings, state)
    print("Open this URL in your browser and approve access:")
    print(auth_url)
    print()
    print(f"Waiting for LinkedIn callback on {settings.linkedin_redirect_uri} ...")

    server = HTTPServer(("localhost", port), CallbackHandler)
    server.handle_request()

    if result.get("error"):
        raise SystemExit(result["error"])
    if not result.get("code"):
        raise SystemExit("LinkedIn callback did not include an authorization code.")

    access_token, person_urn = exchange_raw_code_for_token(settings, result["code"])
    print(f"LINKEDIN_ACCESS_TOKEN={access_token}")
    print(f"LINKEDIN_PERSON_URN={person_urn}")


def fetch_person_urn(access_token: str) -> str:
    response = get_session().get(
        LINKEDIN_USERINFO_URL,
        headers={"Authorization": f"Bearer {access_token}"},
        timeout=30,
    )
    response.raise_for_status()
    subject = response.json()["sub"]
    return f"urn:li:person:{subject}"


def load_content_rows(settings: Settings) -> list[dict[str, str]]:
    source_url = download_url(require(settings.content_source_url, "CONTENT_SOURCE_URL"))
    response = get_session().get(source_url, timeout=60)
    if response.status_code in {401, 403} and "onedrive" in source_url:
        raise SystemExit(
            "OneDrive blocked the workbook download. Open the Excel file, choose Share, "
            "set access to 'Anyone with the link can view', then paste the new link into CONTENT_SOURCE_URL."
        )
    response.raise_for_status()

    source_type = settings.content_source_type
    if source_type == "auto":
        content_type = response.headers.get("Content-Type", "").lower()
        source_path = urlparse(source_url).path.lower()
        if source_path.endswith(".csv") or "csv" in content_type:
            source_type = "csv"
        else:
            source_type = "xlsx"

    if source_type == "csv":
        text = response.content.decode("utf-8-sig")
        return normalize_rows(csv.DictReader(io.StringIO(text)))
    if source_type == "xlsx":
        workbook = load_workbook(io.BytesIO(response.content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value or "").strip().lower() for value in rows[0]]
        parsed_rows = []
        for row in rows[1:]:
            parsed_rows.append(
                {
                    headers[index]: str(value or "").strip()
                    for index, value in enumerate(row)
                    if index < len(headers) and headers[index]
                }
            )
        return normalize_rows(parsed_rows)

    raise SystemExit("CONTENT_SOURCE_TYPE must be auto, csv, or xlsx.")


def download_url(source_url: str) -> str:
    parsed = urlparse(source_url)
    if parsed.netloc.endswith("docs.google.com") and "/spreadsheets/d/" in parsed.path:
        sheet_id = parsed.path.split("/spreadsheets/d/", 1)[1].split("/", 1)[0]
        params = parse_qs(parsed.query)
        gid = params.get("gid", ["0"])[0]
        return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
    if parsed.netloc.endswith("1drv.ms") or parsed.netloc.endswith("onedrive.live.com"):
        encoded_url = base64.urlsafe_b64encode(source_url.encode("utf-8")).decode("utf-8").rstrip("=")
        return f"https://api.onedrive.com/v1.0/shares/u!{encoded_url}/root/content"
    return source_url


def normalize_rows(rows: Any) -> list[dict[str, str]]:
    normalized = []
    for row in rows:
        # Handle both lowercase and capitalized column names
        title_key = next((k for k in row.keys() if k.lower() == "title"), None)
        desc_key = next((k for k in row.keys() if k.lower() == "description"), None)
        
        if not title_key or not desc_key:
            continue
            
        title = str(row.get(title_key, "")).strip()
        description = str(row.get(desc_key, "")).strip()
        if title and description:
            normalized.append({"title": title, "description": description})
    return normalized


def load_history(settings: Settings) -> dict[str, Any]:
    if not os.path.exists(settings.post_history_file):
        return {"posted": []}
    with open(settings.post_history_file, "r", encoding="utf-8") as file:
        return json.load(file)


def save_history(settings: Settings, history: dict[str, Any]) -> None:
    with open(settings.post_history_file, "w", encoding="utf-8") as file:
        json.dump(history, file, indent=2)


def load_scheduler_state(settings: Settings) -> dict[str, Any]:
    if not os.path.exists(settings.scheduler_state_file):
        return {}
    with open(settings.scheduler_state_file, "r", encoding="utf-8") as file:
        return json.load(file)


def save_scheduler_state(settings: Settings, state: dict[str, Any]) -> None:
    with open(settings.scheduler_state_file, "w", encoding="utf-8") as file:
        json.dump(state, file, indent=2)


def row_key(row: dict[str, str]) -> str:
    content = f"{row['title']}\n{row['description']}"
    return hashlib.sha256(content.encode("utf-8")).hexdigest()


def next_content_row(settings: Settings) -> tuple[dict[str, str], str]:
    rows = load_content_rows(settings)
    if not rows:
        raise SystemExit("No usable rows found. Your source must have title and description columns.")

    history = load_history(settings)
    posted = set(history.get("posted", []))
    for row in rows:
        key = row_key(row)
        if key not in posted:
            return row, key

    raise SystemExit("All rows from the source have already been posted.")


def mark_posted(settings: Settings, key: str) -> None:
    history = load_history(settings)
    posted = history.setdefault("posted", [])
    if key not in posted:
        posted.append(key)
    save_history(settings, history)


def generate_post(settings: Settings, source: dict[str, str] | None = None) -> tuple[str, str]:
    require(settings.gemini_api_key, "GEMINI_API_KEY")

    source_text = ""
    if source:
        source_text = f"""
Use this source content from my planning sheet:
Title: {source["title"]}
Description: {source["description"]}
"""

    prompt = f"""
Create a highly structured professional LinkedIn post for a software engineer and a corresponding image description.

Topic area: {settings.post_topic}
Tone: {settings.post_tone}
Audience: {settings.post_audience}
{source_text}

You must output two sections wrapped in XML-style tags:

<POST_TEXT>
A strong, compelling headline.

2-3 short, punchy paragraphs of practical technical content. Include one specific lesson or insight. 

A thoughtful question to encourage engagement.

Exactly 3-5 relevant hashtags.
</POST_TEXT>

<IMAGE_PROMPT>
A detailed, professional visual description for an AI image generator (Stable Diffusion). 
It should be a "Professional tech illustration" or "Clean 3D isometric tech art" that reflects the post topic. 
Avoid text inside the image.
</IMAGE_PROMPT>

Constraint: Post text must be under 1,200 characters. No hype or generic motivational filler.
"""

    url = f"https://generativelanguage.googleapis.com/v1beta/models/{settings.gemini_model}:generateContent"
    headers = {
        "Content-Type": "application/json",
        "X-goog-api-key": settings.gemini_api_key,
    }
    payload = {
        "contents": [
            {
                "parts": [
                    {"text": prompt.strip()}
                ]
            }
        ],
        "generationConfig": {
            "temperature": 0.7,
            "maxOutputTokens": 1024,
        }
    }

    response = get_session().post(url, headers=headers, json=payload, timeout=60)
    response.raise_for_status()
    data = response.json()

    candidates = data.get("candidates", [])
    if not candidates:
        raise SystemExit("Gemini returned no candidates.")

    candidate = candidates[0]
    content = candidate.get("content", {})
    parts = content.get("parts", [])
    
    if parts and isinstance(parts, list):
        full_text = "".join([str(part.get("text", "")) for part in parts]).strip()
        
        post_text = ""
        image_prompt = ""
        
        post_match = re.search(r"<POST_TEXT>(.*?)</POST_TEXT>", full_text, re.DOTALL | re.IGNORECASE)
        image_match = re.search(r"<IMAGE_PROMPT>(.*?)</IMAGE_PROMPT>", full_text, re.DOTALL | re.IGNORECASE)
        
        if post_match:
            post_text = post_match.group(1).strip()
        if image_match:
            image_prompt = image_match.group(1).strip()
            
        if not post_text:
            post_text = full_text # Fallback if tags are missing
        if not image_prompt:
            image_prompt = f"Professional tech illustration: {settings.post_topic}"
            
        return post_text, image_prompt

    raise SystemExit("Unexpected Gemini response format.")


def generate_image(settings: Settings, prompt: str) -> bytes | None:
    """Generate an image using Hugging Face Stable Diffusion API."""
    if not settings.hugging_face_api_key:
        return None
    
    try:
        url = "https://api-inference.huggingface.co/models/stabilityai/stable-diffusion-3.5-large"
        headers = {"Authorization": f"Bearer {settings.hugging_face_api_key}"}
        payload = {"inputs": prompt}
        
        response = get_session().post(url, headers=headers, json=payload, timeout=120)
        response.raise_for_status()
        
        return response.content
    except Exception as e:
        print(f"Warning: Image generation failed: {e}", file=sys.stderr)
        return None


def upload_image_to_linkedin(settings: Settings, image_data: bytes) -> str | None:
    """Upload an image to LinkedIn and return the asset URN."""
    if not image_data:
        return None
    
    try:
        access_token = settings.linkedin_access_token
        person_urn = settings.linkedin_person_urn
        
        # Register the image upload
        register_url = "https://api.linkedin.com/v2/assets?action=registerUpload"
        register_headers = {
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/json",
            "LinkedIn-Version": LINKEDIN_VERSION,
        }
        register_payload = {
            "registerUploadRequest": {
                "owner": person_urn,
                "recipes": ["urn:li:digitalmediaRecipe:feedshare-image"],
                "serviceRelationships": [
                    {
                        "relationshipType": "OWNER",
                        "identifier": "urn:li:userGeneratedContent",
                    }
                ],
            }
        }
        
        register_response = get_session().post(
            register_url, headers=register_headers, json=register_payload, timeout=30
        )
        register_response.raise_for_status()
        register_data = register_response.json()
        
        upload_url = register_data["value"]["uploadMechanism"]["com.linkedin.digitalmedia.uploading.MediaUploadHttpRequest"]["uploadUrl"]
        asset_urn = register_data["value"]["asset"]
        
        # Upload the image
        upload_response = get_session().put(upload_url, data=image_data, timeout=30)
        upload_response.raise_for_status()
        
        return asset_urn
    except Exception as e:
        print(f"Warning: Image upload failed: {e}", file=sys.stderr)
        return None


def publish_post(settings: Settings, text: str, image_urn: str | None = None) -> dict[str, Any]:
    access_token = require(settings.linkedin_access_token, "LINKEDIN_ACCESS_TOKEN")
    author = require(settings.linkedin_person_urn, "LINKEDIN_PERSON_URN")

    payload = {
        "author": author,
        "commentary": text,
        "visibility": settings.post_visibility,
        "distribution": {
            "feedDistribution": "MAIN_FEED",
            "targetEntities": [],
            "thirdPartyDistributionChannels": [],
        },
        "lifecycleState": "PUBLISHED",
        "isReshareDisabledByAuthor": False,
    }
    
    # Add image if available
    if image_urn:
        payload["content"] = {
            "media": {
                "id": image_urn,
            }
        }

    response = get_session().post(
        LINKEDIN_POSTS_URL,
        headers={
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/json",
            "LinkedIn-Version": LINKEDIN_VERSION,
            "X-Restli-Protocol-Version": "2.0.0",
        },
        json=payload,
        timeout=30,
    )
    response.raise_for_status()
    return {"status_code": response.status_code, "post_id": response.headers.get("x-restli-id")}


def preview(settings: Settings) -> None:
    source = None
    if settings.content_source_url and "PASTE_" not in settings.content_source_url:
        source, _ = next_content_row(settings)
        print(f"Source title: {source['title']}")

    text, img_prompt = generate_post(settings, source)
    print("--- POST PREVIEW ---")
    print(text)
    print("\n--- IMAGE PROMPT PREVIEW ---")
    print(img_prompt)


def post(settings: Settings) -> None:
    source = None
    source_key = ""
    if settings.content_source_url and "PASTE_" not in settings.content_source_url:
        source, source_key = next_content_row(settings)
    text, image_prompt = generate_post(settings, source)
    
    image_urn = None
    if settings.hugging_face_api_key:
        image_data = generate_image(settings, image_prompt)
        if image_data:
            image_urn = upload_image_to_linkedin(settings, image_data)
    
    result = publish_post(settings, text, image_urn)
    if source_key:
        mark_posted(settings, source_key)
    print(text)
    print()
    print(f"Published to LinkedIn. Status: {result['status_code']}. Post ID: {result['post_id']}")


def scheduled_post(settings: Settings) -> None:
    today = datetime.now(ZoneInfo("Asia/Kolkata")).date()
    state = load_scheduler_state(settings)
    last_posted_on = parse_date(state.get("last_posted_on", ""))

    # Temporarily disabled interval check for testing/checking
    # if last_posted_on and (today - last_posted_on).days < settings.post_interval_days:
    #     print(f"Skipping. Last successful post was on {last_posted_on.isoformat()}.")
    #     return

    post(settings)
    state["last_posted_on"] = today.isoformat()
    save_scheduler_state(settings, state)


def parse_date(value: str) -> date | None:
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except ValueError:
        return None


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate and publish technical LinkedIn posts.")
    subparsers = parser.add_subparsers(dest="command", required=True)
    subparsers.add_parser("auth-url", help="Print the LinkedIn OAuth authorization URL.")
    subparsers.add_parser("login", help="Start a local callback server and complete LinkedIn OAuth.")
    exchange_parser = subparsers.add_parser("exchange", help="Exchange callback URL for LinkedIn token.")
    exchange_parser.add_argument("callback_url", help="Full redirected callback URL containing ?code=...")
    subparsers.add_parser("preview", help="Generate a post but do not publish it.")
    subparsers.add_parser("post", help="Generate and publish a post.")
    subparsers.add_parser("scheduled-post", help="Post only when the configured interval has passed.")

    args = parser.parse_args()
    settings = load_settings()

    try:
        if args.command == "auth-url":
            print(build_auth_url(settings))
        elif args.command == "login":
            login(settings)
        elif args.command == "exchange":
            access_token, person_urn = exchange_code_for_token(settings, args.callback_url)
            print(f"LINKEDIN_ACCESS_TOKEN={access_token}")
            print(f"LINKEDIN_PERSON_URN={person_urn}")
        elif args.command == "preview":
            preview(settings)
        elif args.command == "post":
            post(settings)
        elif args.command == "scheduled-post":
            scheduled_post(settings)
    except requests.HTTPError as exc:
        response = exc.response
        print(f"HTTP error {response.status_code}: {response.text}", file=sys.stderr)
        raise SystemExit(1) from exc


if __name__ == "__main__":
    main()
